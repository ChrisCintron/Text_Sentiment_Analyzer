#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Word Sentiment Calculator w/ GUI
Authors: Chris Cintron, Matt Novak
Created April 2016


"""
from os import popen
import xlwt
import openpyxl
from tkinter import Button, Label, Grid, Tk, E
from tkinter.filedialog import askopenfilename, asksaveasfile, askdirectory
import re


def openmyfolder():
    global rawInput
    filename = askopenfilename(message= "Pick a .txt file to analyze")
    InputFile = open(filename)
    rawInput = InputFile.read()
    InputFile.close()
    return rawInput

def openmyfolder2():
    filename = askopenfilename(message= "Select the DDods and Warriner Dictionaires")
    global wb1
    wb1 = openpyxl.load_workbook(filename)



def main():

    #Inputs for Original Dict
    #global totalwrdcount
    global uniqwordcount
    global modeValue
    global modeKey


    #Inputs for Dodds
    global freqLabMTWords
    global uniqLabMTWords
    global modeLabMTWordsValue
    global meanLabMTWords
    global modeLabMTWords

    #Inputs for Warriner
    global freqWarriner
    global uniqWarriner
    global modeWarrinerValue
    global meanWarriner
    global modeWarriner


    xlSheet1 = wb1.get_sheet_by_name('labMTwords-English')
    xlSheet2 = wb1.get_sheet_by_name("Warriner-English")

    #import happiness scales as dictionaries
    dictLabMTWords = {}
    dictWarriner = {}

    tempRowNum = xlSheet1.max_row + 1
    for i in range(1, tempRowNum):
        Key = xlSheet1.cell(row=i, column=1).value #cell Ai
        Value = xlSheet1.cell(row=i , column=2).value #cell Bi
        dictLabMTWords.update({ Key : Value }) #Adds them to dictionary

    tempRowNum = xlSheet2.max_row + 1
    for i in range(1, tempRowNum):
        Key = xlSheet2.cell(row=i, column=1).value #cell Ai
        Value = xlSheet2.cell(row=i , column=2).value #cell Bi
        dictWarriner.update({ Key : Value }) #Adds them to dictionary

    #Remove everycharacter EXCEPT a-z, A-Z, and whitespace
    pattern = re.compile('\w+')
    results = re.sub(r'[^a-zA-Z\s]', '', rawInput)
    new_rawInput = results.lower()
    finalInput = pattern.findall(new_rawInput)

    #Loops through each item in list and adds it into dictionary. Adds + 1.
    #If already in dictionary, adds (the value) + 1
    dictInput = {}
    for character in finalInput:
        dictInput.setdefault(character, 0)
        dictInput[character] = dictInput[character] + 1

    uniqwordcount = len(dictInput)
    totalwordcount = 0
    for i in dictInput:
        totalwordcount += dictInput[i]


    #find the mode

    modeKey = []
    modeValue = 0

    for i in dictInput:
        if dictInput[i] == modeValue:
            modeKey.append(i)
        elif dictInput[i] > modeValue:
            modeKey[:] = []
            modeValue = dictInput[i]
            modeKey.append(i)

    #Variables for Dodds
    wordsfoundDodds = {}
    modeLabMTWords = []
    modeLabMTWordsValue = 0
    sumLabMTWords = 0
    freqLabMTWords = 0
    uniqLabMTWords = 0

    for i in dictInput:
        if i in dictLabMTWords.keys():
            Key = i
            Value = dictInput[i]
            wordsfoundDodds.update({ Key : Value }) #Adds them to dictionary

            sumLabMTWords += dictLabMTWords[i] * dictInput[i]
            freqLabMTWords += dictInput[i]
            uniqLabMTWords += 1
            if dictInput[i] == modeLabMTWordsValue:
                modeLabMTWords.append(i)
            elif dictInput[i] > modeLabMTWordsValue:
                modeLabMTWords[:] = []
                modeLabMTWordsValue = dictInput[i]
                modeLabMTWords.append(i)
    meanLabMTWords = sumLabMTWords / freqLabMTWords



    #finds the total score from Warriner scale
    modeWarriner = []
    modeWarrinerValue = 0
    sumWarriner = 0
    freqWarriner = 0
    uniqWarriner = 0

    for i in dictInput:
        if i in dictWarriner.keys():
            sumWarriner += dictWarriner[i] * dictInput[i]
            freqWarriner += dictInput[i]
            uniqWarriner += 1
            if dictInput[i] == modeWarrinerValue:
                modeWarriner.append(i)
            elif dictInput[i] > modeWarrinerValue:
                modeWarriner[:] = []
                modeWarrinerValue = dictInput[i]
                modeWarriner.append(i)
    meanWarriner = sumWarriner / freqWarriner


    ResultsLabel = Label(root, text="Results", bg="blue", fg="white")
    ResultsLabel.grid(row=5, column=0)

    v51 = Label(root, text="Total", bg="white", fg="black")
    v51.grid(row=5, column=1)
    v52 = Label(root, text="Dodds", bg="white", fg="black")
    v52.grid(row=5, column=2)
    v53 = Label(root, text="Warriner", bg="white", fg="black")
    v53.grid(row=5,column=3)


    #Total Word Count Row
    v60 = Label(root, text="Total Word Count", bg="white", fg="black")
    v60.grid(row=6,sticky=E)
    v61 = Label(root, text= totalwordcount, bg="white", fg="black")
    v61.grid(row=6,column=1)
    v62 = Label(root, text= freqLabMTWords, bg="white", fg="black")
    v62.grid(row=6,column=2)
    v63 = Label(root, text= freqWarriner, bg="white", fg="black")
    v63.grid(row=6,column=3)




    #Total Unique Words Row
    v70 = Label(root, text="Total Unique Words", bg="white", fg="black")
    v70.grid(row=7, sticky=E)
    v71 = Label(root, text= uniqwordcount, bg="white", fg="black")
    v71.grid(row=7, column=1)
    v72 = Label(root, text= uniqLabMTWords, bg="white", fg="black")
    v72.grid(row=7, column=2)
    v73 = Label(root, text= uniqWarriner, bg="white", fg="black")
    v73.grid(row=7, column=3)

    #Mode Row
    v80 = Label(root, text="Mode", bg="white", fg="black")
    v80.grid(row=8, sticky=E)
    v81 = Label(root, text= modeValue, bg="white", fg="black")
    v81.grid(row=8, column=1)
    v82 = Label(root, text= modeLabMTWordsValue, bg="white", fg="black")
    v82.grid(row=8, column=2)
    v83 = Label(root, text= modeWarrinerValue, bg="white", fg="black")
    v83.grid(row=8, column=3)

    #Mean Row
    v90 = Label(root, text="Mean", bg="white", fg="black")
    v90.grid(row=9, sticky=E)
    v91 = Label(root, text="N/a", bg="white", fg="black")
    v91.grid(row=9, column=1)
    v92 = Label(root, text= meanLabMTWords, bg="white", fg="black")
    v92.grid(row=9, column=2)
    v93 = Label(root, text= meanWarriner, bg="white", fg="black")
    v93.grid(row=9, column=3)


    #Mode Word
    v10 = Label(root, text="Mode Word", bg="white", fg="black")
    v10.grid(row=10, sticky=E)
    v11 = Label(root, text= modeKey, bg="white", fg="black")
    v11.grid(row=10, column=1)
    v12 = Label(root, text= modeLabMTWords, bg="white", fg="black")
    v12.grid(row=10, column=2)
    v13 = Label(root, text= modeWarriner, bg="white", fg="black")
    v13.grid(row=10, column=3)

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Results')


    #Dictionaries into excel
    rownum = 2
    for i in dictInput:
        ws.write(rownum, 0, i)
        ws.write(rownum, 1, dictInput[i])
        rownum += 1

    ### Writes to Excel, probably will delete ###
    #Headers
    ws.write(0, 0, "Total Words")
    ws.write(0, 1, "Frequency")
    ws.write(0, 3, "Orphan Words")

    #Table for Key Indicator
    #Rows
    ws.write(0, 10, "Total")
    ws.write(0, 11, "Dodds")
    ws.write(0, 12, "Warriner")

    #Collumns
    ws.write(1, 9, "Total Word Count")
    ws.write(2, 9, "Total Unique word count")
    ws.write(3, 9, "Mode")
    ws.write(4, 9, "Mean")
    ws.write(4, 10, "N/a")
    ws.write(5, 9, "Mode Word(s)")

    #Inputs for Original Dict

    ws.write(1, 10, totalwordcount)
    ws.write(2, 10, uniqwordcount)
    ws.write(3, 10, modeValue)
    ws.write(5, 10, modeKey)


    #inputs for Dodds
    ws.write(1, 11, freqLabMTWords)
    ws.write(2, 11, uniqLabMTWords)
    ws.write(3, 11, modeLabMTWordsValue)
    ws.write(4, 11, meanLabMTWords)
    ws.write(5, 11, str(modeLabMTWords))

    #inputs for Warriner
    ws.write(1, 12, freqWarriner)
    ws.write(2, 12, uniqWarriner)
    ws.write(3, 12, modeWarrinerValue)
    ws.write(4, 12, meanWarriner)
    ws.write(5, 12, str(modeWarriner))
    
    #Save into selected directory
    your_dir = askdirectory(message= "Where would you like to save this?")
    wb.save(str(your_dir) + "/word_sentiment_results.xls")


#Window size and title
root=Tk()
#root.configure(background='silver')
root.geometry("480x250+400+200")
root.title("Word Sentiment Calculator")


#Buttons
readbutton = Button(root, text="Open .txt file", bg="white", fg="black", command= openmyfolder)
readbutton.grid(row=1) #Used to find .txt file

readbutton = Button(root, text="Open Excel dictionary", bg="white", fg="black", command= openmyfolder2)
readbutton.grid(row=1, column=2) #Used to find amels_dictionaries


runbutton =  Button(root, text="Run", bg="white", fg="red", command= main)
runbutton.grid(row=1, column=3) #Invokes main() function to run code


#Random Labels
ResultsLabel = Label(root, text="Results", bg="blue", fg="white")
ResultsLabel.grid(row=5, column=0)


root.mainloop()
#HEllO illegalchr branch
